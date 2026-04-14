"""
readers/excel_reader.py
=======================

Читает Excel-файлы (.xlsx / .xls / .xlsm) через openpyxl.
Возвращает SheetData — унифицированный формат для loader.

"""

from __future__ import annotations

import io
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl

from ..exceptions import FileReadError, HeaderValidationError
from .headers import read_headers_raw, validate_headers


@dataclass(frozen=True)
class ExcelReadConfig:
    """Параметры одного чтения Excel-файла.


    skip_header_validation: когда True, заголовки приводятся к нижнему
        регистру и обрезаются, но НЕ проверяются на допустимые символы.
        Используется для шаблонных файлов, где заголовочная строка
        содержит русские отображаемые имена — технические EN-имена
        берутся из TemplateConfig, а не из листа.

    stream: содержимое файла в памяти. Если задано — openpyxl читает из
        BytesIO, не обращаясь к диску. path при этом используется только
        для сообщений об ошибках.
    """

    path: Path
    stream: bytes | None = None
    sheet_name: str | None = None
    skip_rows: int = 0
    skip_cols: int = 0
    max_row: int | None = None
    skip_header_validation: bool = False


@dataclass
class SheetData:
    """Результат read_excel(): заголовки + ленивый итератор строк.

    rows — генератор: книга остаётся открытой до его исчерпания
    или сборки мусора. Не закрывай файл извне.
    """

    headers: list[str]
    rows: Iterator[tuple]
    source_path: Path = field(default_factory=lambda: Path())


# ── Публичный интерфейс ───────────────────────────────────────────────────────

def read_excel(config: ExcelReadConfig) -> SheetData:
    """Открыть Excel-файл, проверить заголовок и вернуть SheetData.

    Книга остаётся открытой на время жизни возвращаемого генератора.
    Закрывается автоматически когда генератор исчерпан или собран GC.

    Args:
        config: параметры чтения (путь, лист, отступы, лимит строк).

    Returns:
        SheetData с заголовками и ленивым итератором строк.

    Raises:
        FileReadError: файл не найден или не читается.
        HeaderValidationError: заголовок некорректен (если валидация включена).
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    source = io.BytesIO(config.stream) if config.stream is not None else config.path
    try:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    except FileNotFoundError:
        raise FileReadError(f"file not found: {config.path}")
    except Exception as exc:
        raise FileReadError(f"cannot open file '{config.path}': {exc}") from exc

    try:
        sheet = wb[config.sheet_name] if config.sheet_name else wb.active
    except KeyError:
        available = wb.sheetnames
        wb.close()
        raise FileReadError(
            f"sheet '{config.sheet_name}' not found. "
            f"Available sheets: {available}"
        )

    header_row_num = config.skip_rows + 1
    try:
        header_raw = next(
            sheet.iter_rows(
                min_row=header_row_num,
                max_row=header_row_num,
                min_col=config.skip_cols + 1,
                values_only=True,
            )
        )
    except StopIteration:
        wb.close()
        raise FileReadError(
            f"sheet '{sheet.title}' has no rows at skip_rows={config.skip_rows}"
        )

    try:
        if config.skip_header_validation:
            headers = read_headers_raw(list(header_raw))
        else:
            headers = validate_headers(list(header_raw))
    except HeaderValidationError:
        wb.close()
        raise

    def _iter_rows() -> Iterator[tuple]:
        try:
            for row in sheet.iter_rows(
                min_row=config.skip_rows + 2,
                min_col=config.skip_cols + 1,
                max_col=config.skip_cols + len(headers),
                max_row=config.max_row,
                values_only=True,
            ):
                if any(cell is not None for cell in row):
                    yield row
        finally:
            wb.close()

    return SheetData(
        headers=headers,
        rows=_iter_rows(),
        source_path=config.path,
    )
# src/manual_excel_loader/template.py
"""
template.py
===========
Парсер Excel-шаблонов со структурой листов 'data' + 'klad_config'.

Публичный API
-------------
read_template_config(path, stream=None) -> TemplateConfig
    Разбирает лист klad_config и возвращает заполненный TemplateConfig.
    Бросает TemplateError при любой структурной ошибке.

is_template(path, stream=None) -> bool
    Быстрая проверка: есть ли в книге оба обязательных листа?

TemplateConfig передаётся в load_template() — тонкую обёртку над load(),
которая предзаполняет LoaderConfig из метаданных шаблона.
"""
from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from .exceptions import FileReadError, TemplateError


# ── Data model ───────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class TemplateConfig:
    """
    Всё что нужно загрузчику, извлечённое из klad_config.

    Атрибуты:
        skip_rows:       Количество строк до строки заголовка на листе 'data'.
                         Вычисляется из номера строки в ячейке B1 klad_config.
        headers:         Технические (EN) имена колонок в порядке klad_config.
                         Становятся именами колонок в SQL/CSV на выходе.
        dtypes:          Имя колонки → строка типа GP.
                         Передаётся напрямую в LoaderConfig.dtypes.
        key_columns:     Множество имён колонок, где NULL недопустим.
        fixed_values:    Имя колонки → литеральное строковое значение для колонок
                         с фиксированным значением (не читается из строк данных).
                         Вставляется в каждую выходную строку.
        russian_headers: Русские отображаемые имена из колонки A klad_config,
                         используются только для проверки совпадения заголовков.
    """
    skip_rows: int
    headers: list[str]
    dtypes: dict[str, str]
    key_columns: frozenset[str]
    fixed_values: dict[str, str]
    russian_headers: list[str]


# ── Internal helpers ─────────────────────────────────────────────────────────

_CELL_ADDR = re.compile(r"^[A-Z]+(\d+)$")


def _parse_skip_rows(cell_value: str) -> int:
    """
    Извлекает skip_rows из ячейки B1 klad_config, например 'A3' → skip_rows=1.

    Ячейка содержит Excel-адрес вида 'A3', означающий что данные начинаются со строки 3.
    Буква колонки значения не имеет — важен только номер строки.
    skip_rows = row_number - 2  (минус 1 за строку заголовка, минус 1 за индексацию с 1)
    """
    if not isinstance(cell_value, str):
        raise TemplateError(
            f"klad_config cell B1 must be an Excel address like 'A3', "
            f"got: {cell_value!r}"
        )
    m = _CELL_ADDR.fullmatch(cell_value.strip().upper())
    if not m:
        raise TemplateError(
            f"klad_config cell B1 must match pattern like 'A3' "
            f"(uppercase letters + digits), got: {cell_value!r}"
        )
    row_number = int(m.group(1))
    if row_number < 2:
        raise TemplateError(
            f"Data cannot start before row 2 (got row {row_number} from '{cell_value}')."
        )
    return row_number - 2 


def _validate_header_alignment(
    data_headers_ru: list[str],
    config_headers_ru: list[str],
    fixed_cols: list[str],
) -> None:
    """
    Проверяет совпадение русских заголовков листа 'data' с заголовками klad_config,
    учитывая что колонки с фиксированным значением на листе 'data' отсутствуют.
    """
    expected = [h for h in config_headers_ru if h not in fixed_cols]

    if data_headers_ru != expected:
        lines = ["Header mismatch between 'data' sheet and 'klad_config':"]
        lines.append(f"  data sheet:   {data_headers_ru}")
        lines.append(f"  klad_config:  {expected}")
        lines.append(
            "  Tip: copy-paste column names from one sheet to the other — "
            "invisible whitespace differences are a common cause."
        )
        raise TemplateError("\n".join(lines))


# ── Public API ───────────────────────────────────────────────────────────────

def is_template(path: Path, stream: bytes | None = None) -> bool:
    """
    Возвращает True если книга содержит оба листа: 'data' и 'klad_config'.

    Если передан stream — читает из него; path используется только в сообщениях об ошибках.
    Содержимое не проверяется — для этого используйте read_template_config().
    """
    import io as _io
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    try:
        source = _io.BytesIO(stream) if stream is not None else path
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        result = "data" in wb.sheetnames and "klad_config" in wb.sheetnames
        wb.close()
        return result
    except Exception:
        return False


def read_template_config(path: Path, stream: bytes | None = None) -> TemplateConfig:
    """
    Разбирает лист 'klad_config' и возвращает TemplateConfig.

    Структура листа klad_config (строки с индексацией с 1):
        Строка 1: B1 = Excel-адрес первой строки данных, например "A3"
        Строка 2: строка заголовков самого klad_config (пропускается)
        Строка 3+: по одной строке на каждую выходную колонку:
            A = русское отображаемое имя (должно совпадать с заголовком листа 'data')
            B = "table" если значение берётся из строк данных,
                или Excel-адрес ячейки (например "A2") для фиксированного значения
            C = "true" если это ключевая колонка (NULL недопустим)
            D = техническое (EN) имя колонки → используется в SQL/CSV на выходе
            E = строка типа GP (например "integer", "text", "timestamp")
        Разбор останавливается на первой полностью пустой строке или sentinel-значении.

    Аргументы:
        path:   путь к файлу Excel-шаблона; если передан stream — используется
                только в сообщениях об ошибках.
        stream: байты уже прочитанного файла (например, скачанного с SMB-шары).
                Если передан — файл читается из него, обращения к диску нет.

    Возвращает:
        TemplateConfig со всеми метаданными, необходимыми загрузчику.

    Бросает:
        FileReadError:  если файл не удаётся открыть.
        TemplateError:  если структура шаблона некорректна (из _parse_klad_config).
    """
    import io as _io
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    source = _io.BytesIO(stream) if stream is not None else path
    try:
        wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
    except FileNotFoundError:
        raise FileReadError(f"file not found: {path}")
    except Exception as exc:
        raise FileReadError(f"cannot open file '{path}': {exc}") from exc

    cfg_sheet = wb["klad_config"]
    data_sheet = wb["data"]

    try:
        result = _parse_klad_config(cfg_sheet, data_sheet)
    finally:
        wb.close()

    return result


_SENTINEL = "<< конец описания шаблона пустая строка в таблице"


def _parse_klad_config(cfg_sheet, data_sheet) -> TemplateConfig:
    """Основная логика разбора — вынесена отдельно для тестируемости."""

    rows = list(cfg_sheet.iter_rows(values_only=True, max_col=6))

    if not rows:
        raise TemplateError("klad_config sheet is empty.")

    # ── Строка 0: B1 содержит адрес первой строки данных ────────────────────
    first_row = rows[0]
    skip_rows = _parse_skip_rows(first_row[1])  # колонка B (0-индекс = 1)

    # ── Строки 2+: определения колонок ───────────────────────────────────────
    russian_headers_cfg: list[str] = []   # колонка A — для проверки заголовков
    tech_headers: list[str] = []          # колонка D — технические имена
    dtypes: dict[str, str] = {}           # tech_name → тип GP
    key_columns: set[str] = set()
    fixed_values: dict[str, str] = {}     # tech_name → фиксированное значение
    fixed_russian_names: list[str] = []   # русские имена колонок с фикс. значением

    for row_idx, row in enumerate(rows[2:], start=2):
        if all(cell is None for cell in row):
            break
        if any(isinstance(cell, str) and _SENTINEL in cell for cell in row):
            break

        ru_name = row[0]   # A: русское отображаемое имя
        source = row[1]    # B: "table" или адрес ячейки
        is_key = row[2]    # C: "true" если ключевая колонка
        tech_name = row[3] # D: техническое EN-имя
        dtype = row[4]     # E: тип GP

        if not isinstance(tech_name, str) or not tech_name.strip():
            raise TemplateError(
                f"klad_config row {row_idx + 1}: column D (technical name) is empty."
            )
        if not isinstance(dtype, str) or not dtype.strip():
            raise TemplateError(
                f"klad_config row {row_idx + 1}: column E (data type) is empty "
                f"for column '{tech_name}'."
            )
        if not isinstance(source, str) or not source.strip():
            raise TemplateError(
                f"klad_config row {row_idx + 1}: column B (source) is empty "
                f"for column '{tech_name}'. Expected 'table' or a cell address."
            )

        tech = tech_name.strip().lower()
        russian_headers_cfg.append(str(ru_name).strip() if ru_name is not None else "")
        tech_headers.append(tech)
        dtypes[tech] = dtype.strip().lower()

        if str(is_key).strip().lower() == "true":
            key_columns.add(tech)

        if source.strip().lower() != "table":
            # фиксированное значение: source — адрес ячейки на листе 'data'
            cell_addr = source.strip().upper()
            if not _CELL_ADDR.fullmatch(cell_addr):
                raise TemplateError(
                    f"klad_config row {row_idx + 1}: column B must be 'table' or "
                    f"a cell address like 'A2', got: {source!r}"
                )
            cell_value = data_sheet[cell_addr].value
            fixed_values[tech] = str(cell_value) if cell_value is not None else ""
            fixed_russian_names.append(str(ru_name).strip() if ru_name is not None else "")

    if not tech_headers:
        raise TemplateError("klad_config defines no columns.")

    # ── Проверка совпадения заголовков с листом 'data' ───────────────────────
    header_row_num = skip_rows + 1  # skip_rows + 1 даёт строку заголовка (1-индекс)
    data_header_row = list(data_sheet.iter_rows(
        min_row=header_row_num,
        max_row=header_row_num,
        values_only=True,
    ))
    if not data_header_row or not data_header_row[0]:
        raise TemplateError(
            f"'data' sheet has no header row at row {header_row_num} "
            f"(derived from skip_rows={skip_rows})."
        )

    data_headers_ru = [
        str(v).strip()
        for v in data_header_row[0]
        if v is not None
    ]
    _validate_header_alignment(data_headers_ru, russian_headers_cfg, fixed_russian_names)

    return TemplateConfig(
        skip_rows=skip_rows,
        headers=tech_headers,
        dtypes=dtypes,
        key_columns=frozenset(key_columns),
        fixed_values=fixed_values,
        russian_headers=russian_headers_cfg,
    )
from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl

from .exceptions import FileReadError, HeaderValidationError


@dataclass(frozen=True)
class ExcelReadConfig:
    """
    Parameters for a single Excel read operation.

    Frozen so that config cannot be accidentally mutated mid-read,
    which would make generator behaviour unpredictable.

    skip_header_validation: when True, header cells are lowercased/stripped
        but NOT validated against the Latin-only regex. Used for template
        files where the header row contains Russian display names — the
        actual column names come from TemplateConfig, not from the sheet.
    """
    path: Path
    sheet_name: str | None = None
    skip_rows: int = 0
    skip_cols: int = 0
    max_row: int | None = None
    skip_header_validation: bool = False


@dataclass
class SheetData:
    """
    Result of read_excel(): validated headers + lazy row iterator.

    rows is a generator — the workbook stays open until the generator
    is exhausted or garbage-collected. Do not close the file externally.
    """
    headers: list[str]
    rows: Iterator[tuple]


# ── Internal helpers ───────────────────────────────────────────────────────────

_VALID_HEADER = re.compile(r"^[a-z0-9_]+$")


def _read_headers_raw(raw: list) -> list[str]:
    """
    Normalise a raw header row without validating character set.

    Used for template files where headers are Russian display names.
    These names are never used as output column names — TemplateConfig
    provides the technical EN names — but we still need to read the row
    to correctly count data rows.
    """
    last_non_none = max(
        (i for i, v in enumerate(raw) if v is not None),
        default=-1,
    )
    if last_non_none == -1:
        raise HeaderValidationError("header row is empty or all cells are None")
    return [str(v).lower().strip() for v in raw[: last_non_none + 1]]


def _validate_headers(raw: list) -> list[str]:
    """
    Validate and normalise a raw header row from openpyxl.

    Raises:
        HeaderValidationError: if headers are empty, contain invalid
            characters, or contain duplicates.
    """
    last_non_none = max(
        (i for i, v in enumerate(raw) if v is not None),
        default=-1,
    )
    if last_non_none == -1:
        raise HeaderValidationError("header row is empty or all cells are None")

    headers = [str(v).lower().strip() for v in raw[: last_non_none + 1]]

    for h in headers:
        if not _VALID_HEADER.fullmatch(h):
            raise HeaderValidationError(
                f"column name '{h}' contains invalid characters. "
                "Only lowercase Latin letters, digits and underscores are allowed."
            )

    if len(headers) != len(set(headers)):
        seen: set[str] = set()
        duplicates = [h for h in headers if h in seen or seen.add(h)]  # type: ignore[func-returns-value]
        raise HeaderValidationError(
            f"duplicate column names are not allowed: {duplicates}"
        )

    return headers


# ── Public interface ───────────────────────────────────────────────────────────

def read_excel(config: ExcelReadConfig) -> SheetData:
    """
    Open an Excel file, validate its header row, and return a SheetData
    with a lazy iterator over data rows.

    The workbook is held open for the lifetime of the returned generator.
    It is closed automatically when:
    - the generator is fully exhausted, or
    - the generator is garbage-collected / explicitly closed.

    Args:
        config: read parameters (path, sheet, offsets, row limit).

    Returns:
        SheetData with headers and a lazy row iterator.

    Raises:
        FileReadError: if the file does not exist or cannot be opened.
        HeaderValidationError: if the header row is invalid (when validation enabled).
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    try:
        wb = openpyxl.load_workbook(config.path, data_only=True, read_only=True)
    except FileNotFoundError:
        raise FileReadError(f"file not found: {config.path}")
    except Exception as exc:
        raise FileReadError(f"cannot open file '{config.path}': {exc}") from exc

    try:
        sheet = wb[config.sheet_name] if config.sheet_name else wb.active
    except KeyError:
        wb.close()
        available = wb.sheetnames
        raise FileReadError(
            f"sheet '{config.sheet_name}' not found. "
            f"Available sheets: {available}"
        )

    header_row_num = config.skip_rows + 1
    try:
        header_raw = next(
            sheet.iter_rows(
                min_row=header_row_num,
                max_row=header_row_num,
                min_col=config.skip_cols + 1,
                values_only=True,
            )
        )
    except StopIteration:
        wb.close()
        raise FileReadError(
            f"sheet '{sheet.title}' has no rows at skip_rows={config.skip_rows}"
        )

    try:
        if config.skip_header_validation:
            headers = _read_headers_raw(list(header_raw))
        else:
            headers = _validate_headers(list(header_raw))
    except HeaderValidationError:
        wb.close()
        raise

    def _iter_rows() -> Iterator[tuple]:
        try:
            for row in sheet.iter_rows(
                min_row=config.skip_rows + 2,
                min_col=config.skip_cols + 1,
                max_col=config.skip_cols + len(headers),
                max_row=config.max_row,
                values_only=True,
            ):
                if any(cell is not None for cell in row):
                    yield row
        finally:
            wb.close()

    return SheetData(headers=headers, rows=_iter_rows())
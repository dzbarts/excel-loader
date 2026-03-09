"""
Тест: при ошибке во время записи частично созданный output-файл удаляется.

Почему это важно:
    Если writer упал на середине — файл существует, но неполный.
    Пользователь мог бы считать загрузку успешной и использовать обрезанные данные.
    Loader должен удалять такой файл в блоке except.

Что проверяем:
    1. output-файл создаётся в процессе записи (pre-condition).
    2. После исключения внутри writer.write() файл не существует.
    3. Исключение пробрасывается наверх (loader не глотает ошибку).
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

from manual_excel_loader.enums import DatabaseType, DumpType, ErrorMode
from manual_excel_loader.models import LoaderConfig


def _make_config(input_file: Path, tmp_path: Path) -> LoaderConfig:
    return LoaderConfig(
        input_file=input_file,
        db_type=DatabaseType.GREENPLUM,
        table_name="t",
        scheme_name="s",
        dump_type=DumpType.SQL,
        error_mode=ErrorMode.IGNORE,
    )


def _make_excel(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    path = tmp_path / "input.xlsx"
    wb.save(path)
    return path


class TestOutputFileCleanupOnError:
    def test_partial_file_deleted_when_writer_raises(self, tmp_path: Path):
        """Если writer.write() бросает исключение — output-файл должен быть удалён.

        Имитируем сбой внутри writer: подменяем SqlFileWriter.write() так,
        что он создаёт файл (pre-condition: файл существует до исключения),
        затем бросает RuntimeError.
        """
        from manual_excel_loader import load

        excel_path = _make_excel(tmp_path)
        config = _make_config(excel_path, tmp_path)

        captured_output_path: list[Path] = []

        def _failing_write(headers, rows):
            # Имитируем частичную запись: создаём файл, потом падаем
            output_path = captured_output_path[0]
            output_path.write_text("partial content")
            raise RuntimeError("disk full")

        with patch(
            "manual_excel_loader.loader.SqlFileWriter"
        ) as mock_writer_cls:
            # Перехватываем момент создания writer, чтобы узнать output_path
            original_init = mock_writer_cls.side_effect

            def capture_init(writer_config):
                captured_output_path.append(writer_config.output_path)
                mock_instance = MagicMock()
                mock_instance.write.side_effect = _failing_write
                return mock_instance

            mock_writer_cls.side_effect = capture_init

            with pytest.raises(RuntimeError, match="disk full"):
                load(config)

        # output_path был захвачен — проверяем что файл удалён
        assert len(captured_output_path) == 1, "Writer должен быть создан ровно один раз"
        assert not captured_output_path[0].exists(), (
            f"Output-файл должен быть удалён после ошибки, "
            f"но существует: {captured_output_path[0]}"
        )

    def test_exception_is_reraised_after_cleanup(self, tmp_path: Path):
        """Loader не должен глотать исключение — оно пробрасывается наверх."""
        from manual_excel_loader import load

        excel_path = _make_excel(tmp_path)
        config = _make_config(excel_path, tmp_path)

        with patch("manual_excel_loader.loader.SqlFileWriter") as mock_writer_cls:
            mock_instance = MagicMock()
            mock_instance.write.side_effect = OSError("no space left on device")
            mock_writer_cls.return_value = mock_instance

            with pytest.raises(OSError, match="no space left"):
                load(config)

    def test_no_file_created_if_writer_raises_before_writing(self, tmp_path: Path):
        """Если writer упал до создания файла — нет файла, нет ошибки удаления."""
        from manual_excel_loader import load

        excel_path = _make_excel(tmp_path)
        config = _make_config(excel_path, tmp_path)

        with patch("manual_excel_loader.loader.SqlFileWriter") as mock_writer_cls:
            mock_instance = MagicMock()
            # writer.write() бросает сразу, не создавая файл
            mock_instance.write.side_effect = PermissionError("access denied")
            mock_writer_cls.return_value = mock_instance

            with pytest.raises(PermissionError):
                load(config)

        # Файлов в tmp_path не должно быть (кроме input.xlsx)
        output_files = [
            f for f in tmp_path.iterdir()
            if f.suffix in (".sql", ".csv") and f.stem != "input"
        ]
        assert output_files == [], f"Не должно быть output-файлов: {output_files}"


class TestValidationReport:
    """Тесты интеграции load() с validation_report: файл отчёта и error_file."""

    def _make_excel_with_bad_data(self, tmp_path: Path) -> Path:
        """Excel с заголовком 'amount' и строками, где значение — строка вместо integer."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["amount"])
        ws.append(["not_a_number"])
        ws.append(["also_bad"])
        path = tmp_path / "data.xlsx"
        wb.save(path)
        return path

    def _make_config_with_validation(
        self,
        input_file: Path,
        report_dir: Path | None = None,
        include_values: bool = False,
        error_mode: "ErrorMode" = None,
    ) -> "LoaderConfig":
        from manual_excel_loader.enums import DatabaseType, DumpType, ErrorMode
        return LoaderConfig(
            input_file=input_file,
            db_type=DatabaseType.GREENPLUM,
            table_name="t",
            scheme_name="s",
            dump_type=DumpType.SQL,
            error_mode=error_mode or ErrorMode.COERCE,
            dtypes={"amount": "integer"},
            validation_report_dir=report_dir,
            validation_report_include_values=include_values,
        )

    def test_no_report_file_when_dir_none(self, tmp_path: Path):
        """validation_report_dir=None → error_file is None, файл не создаётся."""
        from manual_excel_loader import load

        excel = self._make_excel_with_bad_data(tmp_path)
        config = self._make_config_with_validation(excel, report_dir=None)
        result = load(config)

        assert result.error_file is None
        txt_files = list(tmp_path.glob("*_validation_*.txt"))
        assert txt_files == []

    def test_report_file_written_when_dir_set(self, tmp_path: Path):
        """validation_report_dir задан → error_file указывает на созданный TXT."""
        from manual_excel_loader import load

        excel = self._make_excel_with_bad_data(tmp_path)
        report_dir = tmp_path / "reports"
        config = self._make_config_with_validation(excel, report_dir=report_dir)
        result = load(config)

        assert result.error_file is not None
        assert result.error_file.exists()
        assert result.error_file.suffix == ".txt"
        assert "validation" in result.error_file.name

    def test_report_dir_defaults_to_input_parent(self, tmp_path: Path):
        """Если передать input_file.parent — файл ляжет рядом с входным файлом."""
        from manual_excel_loader import load

        excel = self._make_excel_with_bad_data(tmp_path)
        config = self._make_config_with_validation(excel, report_dir=excel.parent)
        result = load(config)

        assert result.error_file is not None
        assert result.error_file.parent == tmp_path

    def test_col_name_in_report(self, tmp_path: Path):
        """Имя колонки из заголовка присутствует в TXT-отчёте."""
        from manual_excel_loader import load

        excel = self._make_excel_with_bad_data(tmp_path)
        report_dir = tmp_path / "reports"
        config = self._make_config_with_validation(excel, report_dir=report_dir)
        result = load(config)

        content = result.error_file.read_text(encoding="utf-8")
        assert "amount" in content

    def test_verify_mode_writes_report_then_raises(self, tmp_path: Path):
        """VERIFY + ошибки + report_dir → файл создаётся, затем DataValidationError."""
        from manual_excel_loader import load
        from manual_excel_loader.enums import ErrorMode
        from manual_excel_loader.exceptions import DataValidationError

        excel = self._make_excel_with_bad_data(tmp_path)
        report_dir = tmp_path / "reports"
        config = self._make_config_with_validation(
            excel, report_dir=report_dir, error_mode=ErrorMode.VERIFY
        )

        with pytest.raises(DataValidationError):
            load(config)

        txt_files = list(report_dir.glob("*_validation_*.txt"))
        assert len(txt_files) == 1, "Файл отчёта должен быть создан до исключения"

    def test_no_report_on_valid_data(self, tmp_path: Path):
        """Если данные валидны — файл отчёта не создаётся (нечего писать)."""
        from manual_excel_loader import load

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["amount"])
        ws.append([42])
        ws.append([7])
        excel = tmp_path / "good.xlsx"
        wb.save(excel)

        report_dir = tmp_path / "reports"
        config = self._make_config_with_validation(excel, report_dir=report_dir)
        result = load(config)

        assert result.error_file is None
        assert not report_dir.exists() or list(report_dir.glob("*.txt")) == []

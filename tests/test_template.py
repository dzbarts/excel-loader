"""
Tests for template.read_template_config() and template.is_template().

We build synthetic workbooks in memory with openpyxl — no fixture files needed.
This makes tests fast, hermetic, and easy to read as living documentation
of the klad_config format.

klad_config sheet layout (1-indexed):
    Row 1: B1 = first-data-row address, e.g. "A3"
    Row 2: header row (skipped)
    Row 3+: col A=ru_name, B=source, C=is_key, D=tech_name, E=dtype
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from src.manual_excel_loader.template import (
    is_template,
    read_template_config,
    TemplateConfig,
    _parse_skip_rows,
)
from src.manual_excel_loader.exceptions import FileReadError, TemplateError


# ── Helpers ──────────────────────────────────────────────────────────────────

def _make_workbook(
    col_defs: list[dict],
    first_data_row: str = "A3",
    data_ru_headers: list[str] | None = None,
    extra_data_rows: list[tuple] | None = None,
    fixed_cell_values: dict[str, str] | None = None,
) -> Path:
    """
    Build a template workbook and return its path (via tmp_path via BytesIO trick).

    col_defs: list of dicts with keys:
        ru_name, source ("table" or cell addr), is_key, tech_name, dtype

    data_ru_headers: Russian names to write on row 1 of 'data' (default: auto from col_defs).
    extra_data_rows: additional data rows to write on 'data' sheet.
    fixed_cell_values: dict of {cell_addr: value} to write on 'data' sheet.
    """
    wb = openpyxl.Workbook()

    # ---- data sheet ----
    ws_data = wb.active
    ws_data.title = "data"

    # Write any fixed-value cells first
    if fixed_cell_values:
        for addr, val in fixed_cell_values.items():
            ws_data[addr] = val

    # Determine the header row number from first_data_row
    import re
    row_num = int(re.search(r"\d+", first_data_row).group())
    header_row = row_num - 1  # one row above data

    # Auto-generate Russian headers from non-fixed col_defs
    if data_ru_headers is None:
        data_ru_headers = [
            d["ru_name"] for d in col_defs if d["source"] == "table"
        ]

    for col_idx, name in enumerate(data_ru_headers, start=1):
        ws_data.cell(row=header_row, column=col_idx, value=name)

    if extra_data_rows:
        for row_offset, data_row in enumerate(extra_data_rows):
            for col_idx, val in enumerate(data_row, start=1):
                ws_data.cell(row=row_num + row_offset, column=col_idx, value=val)

    # ---- klad_config sheet ----
    ws_cfg = wb.create_sheet("klad_config")

    # Row 1: B1 = first_data_row address
    ws_cfg.cell(row=1, column=2, value=first_data_row)

    # Row 2: headers (content doesn't matter, just needs to exist)
    ws_cfg.cell(row=2, column=1, value="Рус. имя")
    ws_cfg.cell(row=2, column=2, value="Источник")
    ws_cfg.cell(row=2, column=3, value="Ключ")
    ws_cfg.cell(row=2, column=4, value="Тех. имя")
    ws_cfg.cell(row=2, column=5, value="Тип")

    # Row 3+: column definitions
    for row_offset, d in enumerate(col_defs):
        r = 3 + row_offset
        ws_cfg.cell(row=r, column=1, value=d.get("ru_name", ""))
        ws_cfg.cell(row=r, column=2, value=d.get("source", "table"))
        ws_cfg.cell(row=r, column=3, value=d.get("is_key", ""))
        ws_cfg.cell(row=r, column=4, value=d.get("tech_name", ""))
        ws_cfg.cell(row=r, column=5, value=d.get("dtype", "text"))

    return wb


def _save_wb(wb, tmp_path: Path, name: str = "template.xlsx") -> Path:
    path = tmp_path / name
    wb.save(path)
    return path


# ── _parse_skip_rows ──────────────────────────────────────────────────────────

class TestParseSkipRows:
    def test_a3_gives_skip_rows_1(self):
        assert _parse_skip_rows("A3") == 1

    def test_a2_gives_skip_rows_0(self):
        assert _parse_skip_rows("A2") == 0

    def test_b5_gives_skip_rows_3(self):
        assert _parse_skip_rows("B5") == 3

    def test_lowercase_accepted(self):
        assert _parse_skip_rows("a3") == 1

    def test_row1_raises(self):
        with pytest.raises(TemplateError, match="row 1"):
            _parse_skip_rows("A1")

    def test_non_address_raises(self):
        with pytest.raises(TemplateError):
            _parse_skip_rows("not_an_address")

    def test_none_raises(self):
        with pytest.raises(TemplateError):
            _parse_skip_rows(None)


# ── is_template ───────────────────────────────────────────────────────────────

class TestIsTemplate:
    def test_both_sheets_present(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "data"
        wb.create_sheet("klad_config")
        path = _save_wb(wb, tmp_path)
        assert is_template(path) is True

    def test_missing_klad_config(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "data"
        path = _save_wb(wb, tmp_path)
        assert is_template(path) is False

    def test_missing_data_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "klad_config"
        path = _save_wb(wb, tmp_path)
        assert is_template(path) is False

    def test_nonexistent_file(self, tmp_path):
        assert is_template(tmp_path / "no_such_file.xlsx") is False


# ── read_template_config — happy path ─────────────────────────────────────────

class TestReadTemplateConfigHappy:
    def _simple_wb(self, tmp_path: Path) -> tuple[Path, list[dict]]:
        col_defs = [
            {"ru_name": "Идентификатор", "source": "table", "is_key": "true",
             "tech_name": "id", "dtype": "integer"},
            {"ru_name": "Наименование", "source": "table", "is_key": "",
             "tech_name": "full_name", "dtype": "text"},
            {"ru_name": "Дата", "source": "table", "is_key": "",
             "tech_name": "hired_at", "dtype": "date"},
        ]
        wb = _make_workbook(col_defs)
        return _save_wb(wb, tmp_path), col_defs

    def test_returns_template_config(self, tmp_path):
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert isinstance(result, TemplateConfig)

    def test_headers_correct(self, tmp_path):
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert result.headers == ["id", "full_name", "hired_at"]

    def test_dtypes_correct(self, tmp_path):
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert result.dtypes == {
            "id": "integer",
            "full_name": "text",
            "hired_at": "date",
        }

    def test_key_columns_detected(self, tmp_path):
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert "id" in result.key_columns
        assert "full_name" not in result.key_columns

    def test_no_fixed_values_by_default(self, tmp_path):
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert result.fixed_values == {}

    def test_skip_rows_from_b1(self, tmp_path):
        # "A3" → data starts at row 3, so skip_rows = 1
        path, _ = self._simple_wb(tmp_path)
        result = read_template_config(path)
        assert result.skip_rows == 1

    def test_multiple_key_columns(self, tmp_path):
        col_defs = [
            {"ru_name": "Год", "source": "table", "is_key": "true",
             "tech_name": "year", "dtype": "integer"},
            {"ru_name": "Месяц", "source": "table", "is_key": "true",
             "tech_name": "month", "dtype": "integer"},
            {"ru_name": "Сумма", "source": "table", "is_key": "",
             "tech_name": "amount", "dtype": "decimal(12,2)"},
        ]
        wb = _make_workbook(col_defs)
        path = _save_wb(wb, tmp_path)
        result = read_template_config(path)
        assert result.key_columns == frozenset({"year", "month"})


class TestReadTemplateConfigFixedValues:
    def test_fixed_value_in_result(self, tmp_path):
        col_defs = [
            {"ru_name": "Источник", "source": "A1", "is_key": "",
             "tech_name": "source_system", "dtype": "text"},
            {"ru_name": "Наименование", "source": "table", "is_key": "",
             "tech_name": "name", "dtype": "text"},
        ]
        wb = _make_workbook(
            col_defs,
            fixed_cell_values={"A1": "ГПН-ИТ"},
        )
        path = _save_wb(wb, tmp_path)
        result = read_template_config(path)
        assert result.fixed_values == {"source_system": "ГПН-ИТ"}

    def test_fixed_col_not_in_table_headers(self, tmp_path):
        """Fixed-value columns should NOT appear in the data sheet header."""
        col_defs = [
            {"ru_name": "Источник", "source": "A1", "is_key": "",
             "tech_name": "source_system", "dtype": "text"},
            {"ru_name": "Имя", "source": "table", "is_key": "",
             "tech_name": "name", "dtype": "text"},
        ]
        wb = _make_workbook(
            col_defs,
            fixed_cell_values={"A1": "ГПН"},
        )
        path = _save_wb(wb, tmp_path)
        result = read_template_config(path)
        # "source_system" must be in headers (output) but not require a data column
        assert "source_system" in result.headers
        assert result.fixed_values["source_system"] == "ГПН"

    def test_none_fixed_cell_becomes_empty_string(self, tmp_path):
        col_defs = [
            {"ru_name": "Пусто", "source": "B1", "is_key": "",
             "tech_name": "empty_col", "dtype": "text"},
            {"ru_name": "Имя", "source": "table", "is_key": "",
             "tech_name": "name", "dtype": "text"},
        ]
        # B1 on data sheet is not set → None → ""
        wb = _make_workbook(col_defs)
        path = _save_wb(wb, tmp_path)
        result = read_template_config(path)
        assert result.fixed_values["empty_col"] == ""


# ── read_template_config — error cases ───────────────────────────────────────

class TestReadTemplateConfigErrors:
    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileReadError):
            read_template_config(tmp_path / "no_such_file.xlsx")

    def test_missing_klad_config_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "data"
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="klad_config"):
            read_template_config(path)

    def test_missing_data_sheet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "klad_config"
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="data"):
            read_template_config(path)

    def test_header_mismatch_raises(self, tmp_path):
        """If Russian names on 'data' don't match klad_config → TemplateError."""
        col_defs = [
            {"ru_name": "Идентификатор", "source": "table", "is_key": "",
             "tech_name": "id", "dtype": "integer"},
        ]
        # Intentionally write a different Russian name to data sheet
        wb = _make_workbook(col_defs, data_ru_headers=["НЕПРАВИЛЬНОЕ ИМЯ"])
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="mismatch"):
            read_template_config(path)

    def test_empty_tech_name_raises(self, tmp_path):
        col_defs = [
            {"ru_name": "Поле", "source": "table", "is_key": "",
             "tech_name": "", "dtype": "text"},
        ]
        wb = _make_workbook(col_defs)
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="technical name"):
            read_template_config(path)

    def test_empty_dtype_raises(self, tmp_path):
        col_defs = [
            {"ru_name": "Поле", "source": "table", "is_key": "",
             "tech_name": "field", "dtype": ""},
        ]
        wb = _make_workbook(col_defs)
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="data type"):
            read_template_config(path)

    def test_invalid_source_address_raises(self, tmp_path):
        col_defs = [
            {"ru_name": "Поле", "source": "NOT_AN_ADDR", "is_key": "",
             "tech_name": "field", "dtype": "text"},
            {"ru_name": "Имя", "source": "table", "is_key": "",
             "tech_name": "name", "dtype": "text"},
        ]
        wb = _make_workbook(col_defs)
        path = _save_wb(wb, tmp_path)
        with pytest.raises(TemplateError, match="cell address"):
            read_template_config(path)
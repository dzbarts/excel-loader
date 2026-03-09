# Стартовый рабочий скрипт
# 

import openpyxl
import csv
from tqdm import tqdm
from datetime import datetime, date, time
import sys
from dateutil import parser
import re
from pathlib import Path
import inspect
from functools import wraps
import warnings
import pandas as pd
# from concurrent.futures import ThreadPoolExecutor # Неоправданно, выигрыш в редких случаях достигает 20%
class RowDataProcessor:
    """
    Класс, позволяющий работать с эксель (csv, sql) файлами для использования в кликхаусе и гринпламе (преимущественно в рамках проектов RHDBI).
    Возможности и соответствующие методы:
        1) выгрузка экселя в sql/csv - process_excel_data()
        2) конвертация sql/csv в эксель - sql_to_excel()/csv_to_excel()
        3) проверка качества данных экселя для GP/CH (в т.ч. шаблонов для загрузки на одс) - process_excel_data()
        4) создание ддл GP/CH - get_ddl()
    Доска в джире: https://kb.gazprom-neft.local/pages/viewpage.action?pageId=526432105
    """
    _DECIMAL_PATTERN = re.compile(r'\(\s*(\d+)\s*,\s*(\d+)\s*\)')
    _CHAR_PATTERN = re.compile(r'\(\s*(\d+)\s*\)')
    _INTERVAL_PATTERN = re.compile(
        r'^'
        r'(?P<value>-?\d+)\s+'
        r'(?P<unit>year|month|day|hour|minute|second)s?'
        r'(\s+(?P<value2>-?\d+)\s+'
        r'(?P<unit2>year|month|day|hour|minute|second)s?)*'
        r'$',
        flags=re.IGNORECASE
    )
    _GP_DTYPE = ['date', 'timestamp', 'real', 'double', 'decimal', 'numeric', 'smallint', 'integer', 'bigint', 'serial', 'bigserial', 'text', 'char', 'array', 'timestamp without time zone', 'boolean', 'tsrange', 'interval']
    _CH_DTYPE = ['Date32', 'DateTime', 'Float64', 'Decimal(32,8)', 'Int16', 'Int32', 'Int64', 'UInt32', 'UInt64', 'String', 'Array', 'DateTime64(6)', 'Bool']
    _ENCODINGS = [
        "utf-8", # Стандарт для современных файлов
        "utf-16", # Unicode (редко используется)
        "utf-16-le", # Little-endian UTF-16
        "utf-16-be", # Big-endian UTF-16
        "utf-32", # Редко используется
        "ascii", # Только латиница (7-bit)
        "latin1", # Западноевропейские языки (iso-8859-1)
        "cp1252", # Западноевропейская (Windows)
        "cp1251", # Кириллица (Windows)
        "cp866", # Кириллица (старые DOS-системы)
        "koi8-r", # Русская кодировка (UNIX)
        "koi8-u", # Украинская кодировка
        "iso-8859-5", # Кириллица (редко используется)
        "gbk", # Китайская (упрощённые иероглифы)
        "big5", # Китайская (традиционные иероглифы)
        "shift_jis", # Японская
        "euc-jp", # Японская (UNIX)
        "euc-kr", # Корейская
    ]
    _DTYPE_RULE = '''
    -- Правила замены типов данных:
    --
    --  GreenPlum:                          ClickHouse:
    -- 'date'                         |     'Date32'
    -- 'timestamp'                    |     'DateTime'
    -- 'real'                         |     'Float64'
    -- 'double'                       |     'Float64'
    -- 'decimal'                      |     'Decimal(32,8)'
    -- 'numeric'                      |     'Decimal(32,8)'
    -- 'smallint'                     |     'Int16'
    -- 'integer'                      |     'Int32'
    -- 'bigint'                       |     'Int64'
    -- 'serial'                       |     'UInt32'
    -- 'bigserial'                    |     'UInt64'
    -- 'text'                         |     'String'
    -- 'char'                         |     'String'
    -- 'array'                        |     'Array'
    -- 'timestamp without time zone'  |     'DateTime64(6)'
    -- 'boolean'                      |     'Bool'
    -- 'double precision'             |     'Int32'
    -- 'character varying'            |     'String'
    -- 'tsrange'                      |     'String'
    -- 'interval'                     |     'String'
    -- 'unknown'                      |     'String'
        '''
    def __init__(self, value=None):
        self._value = value
        self._input_file_path = None
        self._output_file_path = None
        self._error_file_path = None
        self._sheet_name = None
        self._sheet_name_config = None
        self._sheet_data = None
        self._sheet_data_config = None
        self._headers = None
        self._dtypes = None
        self._encoding_input = 'utf-8'
        self._encoding_output = 'utf-8'
        self._db_type = None
        self._set_empty_str_to_null = True
        self._is_strip = False
        self._table_name = 'table_name'
        self._scheme_name = 'scheme_name'
        self._dump_type = 'sql'
        self._delimiter=','
        self._batch_size=10
        self._error='raise'
        self._skip_rows = 0
        self._skip_cols = 0
        self._max_row = None
        self._max_col = None
        self._timestamp = None
        self._timestamp_value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self._extra_template_values = None
        self._wf_load_idn = None
        self._error_message = []
        self._warnings = False
        self._keys = []
        self._keys_idx = []
        self._is_template = False
        self._is_get_ddl = False
    @property
    def __class__(self):
        return RowDataProcessor # для скрытия типа данных результата в _check_type_match
    @staticmethod
    def _clean_column_name(names):
        def simple_transliterate(text):
            cyrillic_map = {
                'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
                'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
                'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
                'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
                'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
                ' ': '_', '-': '_', '_': '_'
            }
            result = []
            for char in text.lower():
                if char in cyrillic_map:
                    result.append(cyrillic_map[char])
                elif char.isalnum():
                    result.append(char)
            return ''.join(result)
        columns = []
        for name in names:
            if not isinstance(name, str):
                return str(name)
            transliterated = simple_transliterate(name)
            cleaned = re.sub(r'[^a-z0-9_]', '', transliterated)
            cleaned = re.sub(r'_{2,}', '_', cleaned.strip('_'))
            columns.append(cleaned)
        result = []
        cnt = {}
        for col in columns:
            if col in cnt:
                cnt[col] += 1
                new_col = f"{col}_{cnt[col]}"
            else:
                cnt[col] = 0
                new_col = col
            result.append(new_col)
        return result
    def get_ddl(self, input_file, db_type, sheet_name=None, skip_rows=None, skip_cols=None, table_name=None, scheme_name=None, timestamp=None, wf_load_idn=None):
        """
        Получение допустимого ддл по экселю.
        Args:
            input_file: полный путь к эксель файлу. Обязательный параметр.
            db_type: тип базы данных. Допустимые значения: 'gp' и 'ch'.
            sheet_name: используемый лист экселя. По умолчанию активный лист.
            skip_rows: кол-во пропускаемых строк в начале экселя.
            skip_cols: кол-во пропускаемых столбцов в начале экселя.
            table_name: наименование таблицы.
            scheme_name: наименование схемы.
            timestamp: временная метка. Допустимые значения: 'write_ts' и 'load_dttm'.
            wf_load_idn: поле-источник для одс.
        Returns: None (принтит ддл)
        """
        self._is_get_ddl = True
        self._get_initiatory(input_file=input_file, sheet_name=sheet_name, db_type=db_type, skip_rows=skip_rows, skip_cols=skip_cols, table_name=table_name, scheme_name=scheme_name, timestamp=timestamp, wf_load_idn=wf_load_idn)
        df = pd.read_excel(self._input_file_path, sheet_name=self._sheet_name, skiprows=self._skip_rows)
        df = df.drop(df.columns[:self._skip_cols], axis=1)
        df_columns_name = self._clean_column_name(list(df.columns))
        df_columns_type = list(df.dtypes)
        df_columns = []
        df_columns_tp = []
        for idx, col in enumerate(df_columns_name):
            if col not in ('write_ts', 'load_dttm'):
                df_columns.append(col)
                df_columns_tp.append(df_columns_type[idx])
            else: self._timestamp = col
        def transform_dtypes(dtype):
            if self._db_type == 'ch':
                match dtype:
                    case 'datetime64[ns]':
                        return 'Date32'
                    case 'timedelta64[ns]':
                        return 'DateTime'
                    case 'float64' | 'float32':
                        return 'Decimal(32,8)'
                    case 'int64' | 'int32':
                        return 'Int32'
                    case 'bool':
                        return 'Bool'
                    case _:
                        return 'String'
            else:
                match dtype:
                    case 'datetime64[ns]':
                        return 'date'
                    case 'timedelta64[ns]':
                        return 'timestamp'
                    case 'float64' | 'float32':
                        return 'decimal'
                    case 'int64' | 'int32':
                        return 'integer'
                    case 'bool':
                        return 'boolean'
                    case _:
                        return 'text'
        columns = []
        for i in range(len(df_columns)):
            if self._db_type == 'ch': columns.append(f"    `{df_columns[i]}` Nullable({transform_dtypes(df_columns_tp[i])}) DEFAULT NULL,\n\n")
            else: columns.append(f"\t{df_columns[i]}\t{transform_dtypes(df_columns_tp[i])}\tNULL,\n")
        if self._db_type == 'ch': columns.append("    `load_dttm` Nullable(DateTime()) DEFAULT NULL\n")
        elif self._db_type == 'gp' and self._timestamp: columns.append(f"\t{self._timestamp}\ttimestamp\tNULL")
        else: columns[-1] = columns[-1].rstrip(",\n")
        ddl_ch = f'''
    -- DROP TABLE IF EXIST {self._scheme_name}.{self._table_name};
    CREATE TABLE {self._scheme_name}.{self._table_name}
    (\n\n{''.join(columns)}\n)
    ENGINE = MergeTree
    ORDER BY load_dttm
    SETTINGS allow_nullable_key = 1,
    index_granularity = 8192;
        '''
        ddl_gp = f'''
    -- {self._scheme_name}.{self._table_name} definition\n
    -- Drop table\n
    -- DROP TABLE {self._scheme_name}.{self._table_name};\n
    CREATE TABLE {self._scheme_name}.{self._table_name}(
    {''.join(columns)}
    )
    DISTRIBUTED RANDOMLY;
        '''
        if self._db_type == 'ch': print(''.join([ddl_ch, '\n', self._DTYPE_RULE]))
        else:
            print(ddl_gp)
            print('Рекомендуется выставлять ключ для распределения вместо randomly')
        if self._scheme_name == 'scheme_name' or self._table_name == 'table_name': print(f'Не забудь заменить scheme_name и\или table_name на свою таблицу приемник')
        self._check_reserved_keywords(df_columns_name)
    def _check_type_match(self, value, expected_type):
        if isinstance(value, type(None)): return RowDataProcessor(value)
        def check_integer(func):
            def wrapped(v, *args, **kwargs):
                try: v = int(v)
                except: return ''
                if not func(v): return f', неудовлетворение ограничению: {extract_lambda_expr(func)}'
                return RowDataProcessor(v)
            return wrapped
        def clear_extra_dtypes_names(x):
            return x[x.find('('):x.find(')')+1]
        def parse_char_params(expected_type):
            match = RowDataProcessor._CHAR_PATTERN.match(clear_extra_dtypes_names(expected_type))
            if match: return int(match.groups(1)[0])
            return None # return size
        def check_decimal(v, expected_type):
            if ',' in str(v): return f', вещественное число должно быть задано через точку'
            try: v=float(v)
            except: return ''
            match = RowDataProcessor._DECIMAL_PATTERN.match(clear_extra_dtypes_names(expected_type))
            if match: precision, scale = map(int, match.groups())
            else: precision, scale = 32, 8
            v = str(v)
            if '.' in v: int_part, dec_part = v.split('.')
            else: int_part, dec_part = v, ''
            int_part = int_part.replace('-', '')
            total_digits = len(int_part) + len(dec_part)
            if total_digits > precision: return f', неудовлетворение ограничению на общее количество символов: {total_digits} > {precision}'
            # if len(dec_part) > scale: return f', неудовлетворение ограничению на количество символов после разделителя: {len(dec_part)} > {scale}'
            if len(dec_part) > scale: v = round(float(v), scale)
            return RowDataProcessor(v)
        def check_char(v, size):
            v = str(v)
            if size:
                if len(v) != size: return f', неудовлетворение ограничению на количество символов в строке: {len(v)} = {size}'
            return RowDataProcessor(v)
        def check_varchar(v, size):
            v = str(v)
            if size:
                if len(v) > size: return f', неудовлетворение ограничению на количество символов в строке: {len(v)} > {size}'
            return RowDataProcessor(v)
        def check_float(func):
            def wrapped(v, *args, **kwargs):
                try: v = float(v)
                except: return ''
                if not func(v): return f', неудовлетворение ограничению: {extract_lambda_expr(func)}'
                return RowDataProcessor(v)
            return wrapped
        def check_datetime(func):
            def wrapped(v, *args, **kwargs):
                try:
                    if not isinstance(v, (datetime, date, time)): v = parser.parse(v)
                except: return ''
                if not func(v): return f', неудовлетворение ограничению: {extract_lambda_expr(func)}'
                if expected_type in ('date', 'date32'): return RowDataProcessor(datetime.strftime(v, '%Y-%m-%d'))
                else: return RowDataProcessor(datetime.strftime(v, '%Y-%m-%d %H:%M:%S'))
            return wrapped
        def check_time(v):
            try:
                if not isinstance(v, (datetime, date, time)): v = parser.parse(v)
            except: return ''
            if isinstance(v, time): return RowDataProcessor(v.strftime('%H:%M:%S'))
            else: return RowDataProcessor(datetime.strftime(v, '%H:%M:%S'))
        def check_array(v):
            if v.strip().startswith('[') and v.strip().endswith(']'): return f', неудовлетворение паттерну типа данных array'
            return RowDataProcessor(v)
        def check_boolean_ch(v):
            v = str(v)
            if v in ('true', 'false'): return RowDataProcessor(v)
            try:
                v = int(v)
                v == 1 or v == 0
            except: return f', неудовлетворение ограничению: значение должно принимать true, false (1, 0)'
            if v == 1: return RowDataProcessor('true')
            else: return RowDataProcessor('false')
        def check_boolean_gp(v):
            v = str(v)
            values = ['true', 'false', '0', '1', 't', 'f', 'y', 'n', 'yes', 'no', 'on', 'off', 'null']
            if v.strip() not in values: return f', значение должно принимать вид из списка {values}'
            return RowDataProcessor(v)
        def check_uuid(v):
            if not bool(re.match(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$', v.strip())): return f', неудовлетворение паттерну типа данных uuid'
            return RowDataProcessor(v)
        def check_interval(v):
            if not bool(RowDataProcessor._INTERVAL_PATTERN.fullmatch(v.strip())): return f', неудовлетворение паттерну типа данных interval'
            return RowDataProcessor(v)
        def check_tsrange(v):
            parts = [part.strip().strip("'") for part in v[1:-1].split(',')]
            start_date, end_date, *rest = parts
            start_date = type_mapping['timestamp'](start_date)
            end_date = type_mapping['timestamp'](end_date)
            if isinstance(start_date, str): return start_date
            else:
                if isinstance(end_date, str): return end_date
                else: return RowDataProcessor(f'{v[0]}{start_date}, {end_date}{v[-1]}') # нет обработки границ
        def extract_lambda_expr(func):
            src = inspect.getsource(func).strip()
            return src.rpartition(':')[2][:-2].replace("x", "значение")
        def check_string(v):
            return RowDataProcessor(v)
        type_mapping_ch = {
            # Целочисленные типы
            'int8': check_integer(lambda x: -128 <= x <= 127),
            'int16': check_integer(lambda x: -32768 <= x <= 32767),
            'int32': check_integer(lambda x: -2147483648 <= x <= 2147483647),
            'int64': check_integer(lambda x: -9223372036854775808 <= x <= 9223372036854775807),
            'int128': check_integer(lambda x: -2**127 <= x <= 2**127),
            'int256': check_integer(lambda x: -2**255 <= x <= 2**255-1),
            'uint8': check_integer(lambda x: 0 <= x <= 255),
            'uint16': check_integer(lambda x: 0 <= x <= 65535),
            'uint32': check_integer(lambda x: 0 <= x <= 4294967295),
            'uint64': check_integer(lambda x: 0 <= x <= 18446744073709551615),
            'uint128': check_integer(lambda x: 0 <= x <= 2**128-1),
            'uint256': check_integer(lambda x: 0 <= x <= 2**256-1),
            # Числа с плавающей точкой
            'float32': check_float(lambda x: -3.4e38 <= x <= 3.4e38),
            'float64': check_float(lambda x: -1.7e308 <= x <= 1.7e308),
            # Числа с фиксированной или произвольной точностью
            # Рассмотрены отдельно
            # Строковые типы
            'string': lambda x: check_string(x),
            # 'FixedString': lambda v, x: len(v)==x,
            # Даты и время
            'date': check_datetime(lambda x: datetime(1970, 1, 1) <= x <= datetime(2105, 12, 31)),
            'date32': check_datetime(lambda x: datetime(1900, 1, 1) <= x <= datetime(2299, 12, 31)),
            'datetime': check_datetime(lambda x: datetime(1970, 1, 1) <= x <= datetime(2105, 12, 31, 23, 59, 59)),
            'datetime64': check_datetime(lambda x: datetime(1900, 1, 1) <= x <= datetime(2299, 12, 31, 23, 59, 59)),
            'datetime32': check_datetime(lambda x: datetime(1900, 1, 1) <= x <= datetime(2299, 12, 31)),
            'time': lambda x: check_time(x),
            # Другие
            'array': lambda x: check_array(x),
            'bool': lambda x: check_boolean_ch(x),
            'uuid': lambda x: check_uuid(x)
        }
        type_mapping_gp = {
            # Целочисленные типы
            'smallint': check_integer(lambda x: -32768 <= x <= 32767),
            'integer': check_integer(lambda x: -2147483648 <= x <= 2147483647),
            'bigint': check_integer(lambda x: -9223372036854775808 <= x <= 9223372036854775807),
            # Числа с плавающей точкой
            'real': check_float(lambda x: -3.4e38 <= x <= 3.4e38),
            'double precision': check_float(lambda x: -1.7e308 <= x <= 1.7e308),
            # Числа с фиксированной или произвольной точностью
            # Рассмотрены отдельно
            # Автоинкрементные числовые типы
            'smallserial': check_integer(lambda x: -32768 <= x <= 32767),
            'serial': lambda x: check_integer(lambda x: 1 <= x <= 2147483647),
            'bigserial': lambda x: check_integer(lambda x: 1 <= x <= 9223372036854775807),
            # Строковые типы
            # char и varchar рассмотрены отдельно
            'text': lambda x: check_string(x),
            # Даты и время
            'date': check_datetime(lambda x: datetime(1, 1, 1) <= x <= datetime(9999, 12, 31)),
            'time': lambda x: check_time(x),
            'timestamp': check_datetime(lambda x: datetime(1, 1, 1) <= x <= datetime(9999, 12, 31, 23, 59, 59)),
            'interval': lambda x: check_interval(x), # нет ограничения на минимальную и максимальную дату
            'tsrange': lambda x: check_tsrange(x),
            # 'smalldatetime':
            # Другие
            'boolean': lambda x: check_boolean_gp(x),
            'uuid': lambda x: check_uuid(x)
            # array будет добавлен позднее
        }
        type_mapping = type_mapping_gp if self._db_type == 'gp' else type_mapping_ch
        try: return type_mapping[expected_type](value)
        except KeyError:
            if 'dec' in expected_type.lower() or 'numeric' in expected_type.lower(): return check_decimal(value, expected_type)
            elif 'varchar' in expected_type.lower(): return check_varchar(value, parse_char_params(expected_type))
            elif 'char' in expected_type.lower(): return check_char(value, parse_char_params(expected_type))
            else: return 'Некорректно обрабатывается тип данных'
    def _uni_col_names(self, columns):
        columns = [x.lower().strip() for x in columns]
        out_columns = []
        if self._db_type == 'ch':
            for col in columns:
                if col in ('tinyint', 'int8'): out_columns.append('int8')
                elif col in ('smalint', 'int16'): out_columns.append('int16')
                elif col in ('int', 'integer', 'int32'): out_columns.append('int32')
                elif col in ('bigint', 'int64'): out_columns.append('int64')
                elif col in ('float', 'float32'): out_columns.append('float32')
                elif col in ('double', 'float64'): out_columns.append('float64')
                elif 'dec' in col: out_columns.append(col)
                elif 'datetime64' in col: out_columns.append('datetime64')
                elif 'datetime32' in col: out_columns.append('datetime32')
                elif 'datetime' in col or 'timestamp' in col: out_columns.append('datetime')
                elif 'date32' in col: out_columns.append('date32')
                elif 'date' in col: out_columns.append('date')
                elif 'bool' in col: out_columns.append('bool')
                elif 'array' in col: out_columns.append('array')
                elif col in ('string', 'uint8', 'uint16', 'uint32', 'uint64', 'uint128', 'uint256', 'int128', 'int256', 'time', 'uuid'): out_columns.append(col)
                else:
                    print(f'Тип данных {col} не найден или не обрабатывает, поле проверено не будет.\nДопустимые типы данных: {self._CH_DTYPE}')
                    out_columns.append('string')
        else:
            for col in columns:
                if col in ('smallint', 'int2'): out_columns.append('smallint')
                elif col in ('integer', 'int', 'int4'): out_columns.append('integer')
                elif col in ('bigint', 'int8'): out_columns.append('bigint')
                elif col in ('smallserial', 'serial2'): out_columns.append('smallserial')
                elif col in ('serial', 'serial4'): out_columns.append('serial')
                elif col in ('bigserial', 'serial8'): out_columns.append('bigserial')
                elif col in ('real', 'float4'): out_columns.append('real')
                elif col in ('double precision', 'double', 'float8', 'float'): out_columns.append('double precision')
                elif 'numeric' in col or 'dec' in col or 'char' in col: out_columns.append(col)
                elif 'timestamp' in col: out_columns.append('timestamp')
                elif 'date' in col: out_columns.append('date')
                elif 'tsrange' in col: out_columns.append('tsrange')
                elif 'bool' in col: out_columns.append('boolean')
                elif 'array' in col: out_columns.append('array')
                elif col in ('text', 'interval', 'uuid', 'time'): out_columns.append(col)
                else:
                    print(f'Тип данных {col} не найден или не обрабатывает, поле проверено не будет.\nДопустимые типы данных: {self._GP_DTYPE}')
                    out_columns.append('text')
        return out_columns
    @staticmethod
    def _get_excel_cell_name(row, column):
        column_letter = ''
        while column > 0:
            column, remainder = divmod(column-1, 26)
            column_letter = chr(65+remainder)+column_letter
        return f'{column_letter}{row}'
    def _check_row_validity(self, row, expected_types, row_idx):
        output_row = []
        for col_idx, (cell, expected_type) in enumerate(zip(row, expected_types)):
            result = self._check_type_match(cell, expected_type)
            if self._keys != [] and col_idx in self._keys_idx and cell is None:
                self._keys[self._keys_idx.index(col_idx)] = False
            if not isinstance(result, str):
                try:
                    output_row.append(result._value)
                except:
                    print(result)
                    sys.exit()
            else:
                output_row.append(None)
                if row_idx is not None: self._error_message.append((self._get_excel_cell_name(row_idx+self._skip_rows+2, col_idx+self._skip_cols+1), cell, f'Ожидаемый тип: {expected_type}{result}'))
        return output_row
    def _extract_types_from_ddl(self, ddl_string):
        if self._db_type == 'ch': return self._extract_types_from_ddl_ch(ddl_string)
        else: return self._extract_types_from_ddl_gp(ddl_string)
    @staticmethod
    def _extract_types_from_ddl_ch(ddl_string):
        processed_ddl = re.sub(r'(\([\s\n]*\d+[\s\n]*,[\s\n]*\d+[\s\n]*\))',
                            lambda m: m.group(1).replace('\n', ''),
                            ddl_string)
        processed_ddl = processed_ddl.replace('\n', ' ')
        match = re.search(r'\((.*)\)', processed_ddl)
        if not match:
            return []
        columns_def = match.group(1)
        columns = []
        current_column = ''
        bracket_count = 0
        for char in columns_def:
            if char == '(':
                bracket_count += 1
            elif char == ')':
                bracket_count -= 1
            elif char == ',' and bracket_count == 0:
                columns.append(current_column.strip())
                current_column = ''
                continue
            current_column += char
        if current_column:
            columns.append(current_column.strip())
        types = []
        for column in columns:
            nullable_decimal_match = re.search(r'\s+Nullable\(Decimal\((\d+),\s*(\d+)\)\)', column)
            if nullable_decimal_match:
                precision, scale = nullable_decimal_match.groups()
                types.append(f'Decimal({precision},{scale})')
                continue
            nullable_match = re.search(r'\s+Nullable\((.*?)\)', column)
            if nullable_match:
                inner_type = nullable_match.group(1)
                types.append(inner_type)
                continue
            decimal_match = re.search(r'\s+Decimal\((\d+),\s*(\d+)\)', column)
            if decimal_match:
                precision, scale = decimal_match.groups()
                types.append(f'Decimal({precision},{scale})')
                continue
            type_match = re.search(r'\s+([A-Za-z]+(?:\([^)]+\))?)', column)
            if type_match:
                types.append(type_match.group(1))
        return types
    @staticmethod
    def _extract_types_from_ddl_gp(ddl_string):
        ddl_clean = re.sub(r'--.*?\n', '', ddl_string)
        ddl_clean = re.sub(r'/\*.*?\*/', '', ddl_clean, flags=re.DOTALL)
        ddl_clean = ' '.join(ddl_clean.split())
        match = re.search(r'CREATE\s+(?:TEMPORARY\s+)?TABLE\s.*?\((.*)\)', ddl_clean, re.IGNORECASE | re.DOTALL)
        if not match:
            return []
        columns_part = match.group(1)
        column_defs = []
        current_def = []
        paren_level = 0
        for token in re.split(r'([(),])', columns_part):
            if token == '(': paren_level += 1
            elif token == ')': paren_level -= 1
            elif token == ',' and paren_level == 0:
                column_defs.append(''.join(current_def).strip())
                current_def = []
                continue
            current_def.append(token)
        if current_def:
            column_defs.append(''.join(current_def).strip())
        types = []
        for col_def in column_defs:
            if not col_def or col_def.startswith(('CONSTRAINT', 'PRIMARY KEY', 'FOREIGN KEY')): continue
            type_match = re.match(
                r'^"?[\w_]+"?\s+('
                r'(?:'
                r'\w+(?:\s*\(\s*\d+\s*(?:,\s*\d+\s*)?\))?' # типы со скобками
                r'(?:\s+with(?:out)?\s+time\s+zone)?' # timestamp with time zone
                r'|'
                r'\w+(?:\s+\w+)+' # составные типы
                r')'
                # r'(?:\s+(?:NOT\s+NULL|NULL|DEFAULT\s+[^,]+))?' # модификаторы
                r')',
                col_def,
                re.IGNORECASE
            )
            if type_match: types.append(type_match.group(1).strip())
        return types
    def _check_headers(self, headers):
        last_non_none_idx = max((i for i , el in enumerate(headers) if el is not None), default=-1)
        if not any(isinstance(header, str) for header in headers): sys.exit(f'Некорректное наименование поля {header} в заголовке {headers}')
        headers = [x.lower().strip() for x in headers[:last_non_none_idx+1]]
        if len(headers) == 0: sys.exit(f'Не удалось найти загловок, было найдено {headers}')
        if self._is_get_ddl is False:
            for header in headers:
                if not isinstance(header, str): sys.exit(f'Некорректное наименование поля {header} в заголовке {headers}')
                if not re.fullmatch(r'^[a-z0-9_]+$', header):
                  print(header)
                  sys.exit(f'Поле {header} содержит недопустимые символы в заголовке {headers}.\n(Допускается использование латиницы, цифр и нижнего подчеркивания)')
                if header.isdigit(): sys.exit(f'Поле {header} содержит только цифры в заголовке {headers}.\n(Допускается использование только цифр, но крайне не рекомендуется.)')
            if len(headers) != len(set(headers)): sys.exit(f'Запрещено одинаковое наименование разных полей в заголовке {headers}.')
        self._headers = headers
        self._max_col = len(self._headers)+self._skip_cols
    def _get_initiatory(self, input_file, sheet_name=None, sheet_name_config=None, output_file=None, error_file=None, dump_type=None, encoding_input=None, encoding_output=None, db_type=None, skip_rows=None, skip_cols=None, dtypes=None, table_name=None, scheme_name=None, batch_size=None, error=None, delimiter=None, timestamp=None, max_row=None, wf_load_idn=None, is_strip=None, set_empty_str_to_null=None):
        if self._warnings is False: warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        if is_strip: self._is_strip = is_strip
        if set_empty_str_to_null: self._set_empty_str_to_null = set_empty_str_to_null
        if encoding_input:
            if encoding_input.strip().lower() in _ENCODINGS: self._encoding_input = encoding_input
            else: sys.exit(f'Кодировка {encoding_input} не поддерживается.\nРекомендуется использовать "utf-8" или "cp-1251".\nМожно также выбрать из списка ниже {_ENCODINGS}. ')
        if encoding_output:
            if encoding_output.strip().lower() in _ENCODINGS: self._encoding_output = encoding_output
            else: sys.exit(f'Кодировка {encoding_output} не поддерживается.\nРекомендуется использовать "utf-8" или "cp-1251".\nМожно также выбрать из списка ниже {_ENCODINGS}. ')
        if dump_type:
            if dump_type.strip().lower() in ('sql', 'csv'): self._dump_type = dump_type
            else: sys.exit(f'Тип исходящего файла может быть "sql" или "csv", однако был получен {dump_type}.')
        self._input_file_path = Path(input_file)
        if sheet_name: self._sheet_name = sheet_name
        print('Excel начал обработку. Может занять немного больше времени, чем предполагается')
        try: workbook = openpyxl.load_workbook(input_file, data_only=True, read_only=False)
        except: sys.exit(f'Невозможно прочитать эксель файл, скорее всего неправильно указан путь {input_file}')
        if not workbook: sys.exit(f'Эксель по пути {input_file} не найден')
        self._is_template = bool('data' in workbook.sheetnames and 'klad_config' in workbook.sheetnames)
        if error:
            if error in ('raise', 'coerce', 'ignore', 'verify'): self._error = error
            else: sys.exit(f'Некорректное наименование типа исследования ошибок {error}, error может принимать значения "raise", "coerce", "ignore" или "verify".')
        if self._is_template:
            self._sheet_name = 'data'
            self._sheet_data = workbook['data']
            self._sheet_data_config = workbook['klad_config']
            if skip_rows: print(f'При загрузке шаблонов невозможно установить параметр skip_rows')
            if skip_cols: print(f'При загрузке шаблонов невозможно установить параметр skip_cols')
            if self._sheet_name: print(f'При загрузке шаблонов невозможно установить параметр sheet_name, инструмент работает с листами "data" и "klad_config"')
            if db_type: print(f'При загрузке шаблонов невозможно установить параметр db_type, инструмент работает с типом базы данных "gp"')
            if dtypes: print(f'При загрузке шаблонов невозможно установить параметр dtypes, инструмент работает с типами данных листа "klad_config"')
        else:
            if skip_rows: self._skip_rows += skip_rows
            if skip_cols: self._skip_cols += skip_cols
            if self._sheet_name: self._sheet_data = workbook[self._sheet_name]
            else:
                self._sheet_data = workbook.active
                self._sheet_name = self._sheet_data.title
            if db_type:
                if db_type.strip().lower() in ('ch', 'gp'): self._db_type = db_type
                else: sys.exit(f'Тип базы данных может быть "ch" или "gp", однако был получен {db_type}.')
            if dtypes:
                if not self._db_type: sys.exit(f'Необходимо определить базу данных: db_type="ch" или db_type="gp".')
                if isinstance(dtypes, str): dtypes = self._extract_types_from_ddl(dtypes)
                self._dtypes = self._uni_col_names(dtypes)
            if self._sheet_data:
                headers = [cell.value for cell in self._sheet_data[self._skip_rows+1][self._skip_cols:]]
                self._check_headers(headers)
            else: sys.exit(f'Не удалось прочитать данные, было получено {self._sheet_data}')
        if output_file: self._output_file_path = Path(output_file).with_suffix(f'.{self._dump_type}')
        else: self._output_file_path = self._input_file_path.with_name(f'{self._input_file_path.stem}_{datetime.now().strftime('%d%m%y_%H%M%S')}').with_suffix(f'.{self._dump_type}')
        if error_file: self._error_file_path = self._error_file_path = Path(error_file)
        else: self._error_file_path= self._input_file_path.with_name(f'error_{self._input_file_path.stem}_{datetime.now().strftime('%d%m%y_%H%M%S')}').with_suffix('.txt')
        if table_name: self._table_name = table_name
        if scheme_name: self._scheme_name = scheme_name
        if batch_size: self._batch_size = batch_size
        if delimiter: self._delimiter = delimiter
        if self._is_template: self._get_template_init(input_file)
        if max_row: self._max_row = max_row + self._skip_rows + 1
        if wf_load_idn:
            self._wf_load_idn = wf_load_idn
            self._headers.append('wf_load_idn')
        if timestamp:
            if timestamp.strip().lower() in ('write_ts', 'load_dttm'):
                if timestamp.strip().lower() not in self._headers:
                    self._timestamp = timestamp
                    self._headers.append(timestamp)
            else: sys.exit(f'Тип временной метки может быть "write_ts" (для ODS) или "load_dttm" (для DM), однако был получен {timestamp}.')
    def _get_template_init(self, input_file):
        dtypes = []
        headers = []
        headers_russ_conf = []
        extra_template_values = []
        for idx, row in enumerate(self._sheet_data_config.iter_rows(values_only=True, max_col=6)):
            if (all(cell is None for cell in row) or any(cell=='<< конец описания шаблона пустая строка в таблице' for cell in row)):
                break
            else:
                if idx == 0:
                    cell = row[1]
                    if not re.fullmatch(r'[A-Z]+[0-9]+', cell): sys.exit(f'В значении должно быть два набора символов: первые - заглавные латиницы, а вторые - цифры, однако получено "{cell}"')
                    else: self._skip_rows += int(re.fullmatch(r'[A-Z]+([0-9]+)', cell).group(1))-2
                    continue
                if idx == 1: continue
                headers.append(row[3])
                headers_russ_conf.append(row[0])
                if row[1] != 'table':
                    cell = row[1]
                    if not re.fullmatch(r'[A-Z]+[0-9]+', cell): sys.exit(f'В значении должно быть два набора символов: первые - заглавные латиницы, а вторые - цифры, однако получено "{cell}"')
                    extra_template_values.append((str(self._sheet_data[cell].value), idx-2))
                if row[1] == 'table':
                    if row[2] == 'true': self._keys_idx.append(idx-2-len(extra_template_values))
                    dtypes.append(row[4])
        headers_russ = [cell.value for cell in self._sheet_data[self._skip_rows+1][:len(headers_russ_conf)-len(extra_template_values)]]
        headers_russ = [str(x).strip() for x in headers_russ] # раньше лишние пробелы давали сбой
        headers_russ_conf = [x.strip() for x in headers_russ] # аналогично
        headers_russ_conf_new = [x for x in headers_russ_conf if x in headers_russ] # нет проверки на наименование полей при фиксированных значениях, пропускаем их
        if headers_russ != headers_russ_conf_new:
            print('Нарушено соответствие наименование полей между листами (поля A)')
            print('(Рекомендуется сделать ctrl+c ctrl+v с одного листа на другой, тк бывает, что визуально они одинаковые, но на деле нет)')
            print('Наименование полей:\t\t(Наименования полея для фикс значений не учитываются)')
            for _ in range(len(extra_template_values)): headers_russ.insert(0, None)
            headers_russ.insert(0, 'Лист data')
            headers_russ_conf.insert(0, 'Лист klad_config')
            for x, y in zip(headers_russ, headers_russ_conf):
                print(f'"{x}"\t\t\t"{y}"')
            sys.exit()
        self._db_type = 'gp'
        self._dtypes = self._uni_col_names(dtypes)
        if headers: self._check_headers(headers)
        self._extra_template_values = extra_template_values
        self._max_col -= len(extra_template_values)
        # self._keys_idx = [x-len(extra_template_values) for x in self._keys_idx]
        self._keys = [True] * len(self._keys_idx)
    def _insert_extra_values(self, row):
        row = list(row)
        if self._extra_template_values:
            for v, i in self._extra_template_values:
                row.insert(i, v)
        if self._wf_load_idn:
            row.append(self._input_file_path.name)
        if self._timestamp:
            row.append(self._timestamp_value)
        return tuple(row)
    @staticmethod
    def _excel_cell_for_sort(cell):
        parts = re.match(r"([A-Z]+)(\d+)", cell)
        if parts: return parts.group(1), int(parts.group(2))
    def _get_error_message(self, message):
        out_message = ''
        error_pr = None
        idx = 0
        for cell_name, cell, error in sorted(message, key=lambda x: (x[2], self._excel_cell_for_sort(x[0]))):
            if error != error_pr:
                if error_pr is not None: out_message += '\n'
                error_pr = error
                idx += 1
                out_message += f"{'='*(len(error)+len(str(idx))+16)}\n"
                out_message += f'{idx}. Тип ошибки: {error}\n'
                out_message += f"{'='*(len(error)+len(str(idx))+16)}\n"
                out_message += f"\tЯчейка\t|\tЗначение\n"
                out_message += f"\t      \t|\t        \n"
            out_message += f"\t{cell_name}\t|\t{cell}\n"
        if self._keys != [] and False in self._keys:
            if error_pr is not None: out_message += '\n'
            out_message += f"{'='*(len(str(idx))+62)}\n"
            out_message += f'{idx+1}. Тип ошибки: ключевое поле не может содержать пустые значения\n'
            out_message += f"{'='*(len(str(idx))+62)}\n"
            out_message += f"\tПоле\t|\tСожержит ли пустое значение\n"
            out_message += f"\t      \t|\t        \n"
            for idx, key in enumerate(self._keys):
                if key is False:
                    out_message += f"\t{self._headers[self._keys_idx[idx]]}\t|\tДа\n"
            if self._extra_template_values:
                for v, i in self._extra_template_values:
                    if v == str(None):
                        out_message += f"\t{self._headers[i]}\t|\tДа\n"
        return out_message
    def process_excel_data(self, input_file, sheet_name=None, output_file=None, dump_type=None, encoding_input=None, encoding_output=None, db_type=None, skip_rows=None, skip_cols=None, dtypes=None, table_name=None, scheme_name=None, batch_size=None, error=None, delimiter=None, timestamp=None, max_row=None, wf_load_idn=None, is_strip=None, set_empty_str_to_null=None):
        """
        Опционально до двух действий: валидация данных и выгрузка в sql/csv.
        Args:
            input_file: полный путь к эксель файлу. Обязательный параметр.
            sheet_name: используемый лист экселя. По умолчанию активный лист.
            output_file: полный путь к исходящему sql/csv файлу. По умолчанию название входящего файла + метка времени.
            dump_type: тип исходящего файла. Допустимые значения: 'sql' и 'csv'. По умолчанию 'sql'.
            encoding_input: кодировка экселя. По умолчанию 'utf-8'.
            encoding_output: кодировка исходящего файла. По умолчанию 'utf-8'.
            db_type: тип базы данных. Допустимые значения: 'gp' и 'ch'.
            skip_rows: кол-во пропускаемых строк в начале экселя.
            skip_cols: кол-во пропускаемых столбцов в начале экселя.
            dtypes: типы данных для валидации или ддл.
            table_name: наименование таблицы для исходящего sql файла.
            scheme_name: наименование схемы для исходящего sql файла.
            batch_size: размер батча для исходящего sql файла.
            error: тип обработки ошибок в эксель файле. Допустимые значения: 'verify', 'raise', 'ignore' и 'coerce'. По умолчанию 'raise'.
                ignore: выгрузка данных как есть;
                coerce: выгрузка данных с заменной всех невалидных данных на NULL;
                verify: проверка качества данных;
                raise: проверка качества данных и выгрузка данных с заменной всех невалидных данных на NULL;
            delimiter: тип разделителя для исходящего csv файла. По умолчанию запятая.
            timestamp: временная метка для исходящего файла. Допустимые значения: 'write_ts' и 'load_dttm'.
            max_row: кол-во обрабатываемых строк.
            wf_load_idn: поле-источник для исходящего файла для одс.
            is_strip: обрезание пустых символов в начале и в конце ячеек экселя. Допустимые значения: True и False. По умолчанию False.
            set_empty_str_to_null: преобразование пустых ячеек экселя в NULL. Допустимые значения: True и False. По умолчанию True.
        Returns: None (создает файл с ошибками и/или файл sql/csv)
        """
        print('Не взаимодействуйте со входящими и исходящими файлами во избежание некорректной работы инструмента')
        self._get_initiatory(input_file=input_file, sheet_name=sheet_name, output_file=output_file, dump_type=dump_type, encoding_input=encoding_input, encoding_output=encoding_output, db_type=db_type, skip_rows=skip_rows, skip_cols=skip_cols, dtypes=dtypes, table_name=table_name, scheme_name=scheme_name, batch_size=batch_size, error=error, delimiter=delimiter, timestamp=timestamp, max_row=max_row, wf_load_idn=wf_load_idn, is_strip=is_strip, set_empty_str_to_null=set_empty_str_to_null)
        if self._error in ('verify', 'raise') and not self._dtypes: sys.exit(f'Для проверки качества данных необходимо указать типы данных (параметр dtype).')
        if error in ('verify', 'raise'):
            with open(self._error_file_path, mode='w', newline='', encoding=self._encoding_output) as error_file:
                for row_idx, row in tqdm(enumerate(self._sheet_data.iter_rows(min_row=self._skip_rows+2, min_col=self._skip_cols+1, values_only=True, max_row=self._max_row, max_col=self._max_col))):
                    if any(cell is not None for cell in row):
                        self._check_row_validity(row, self._dtypes, row_idx)
                error_file.write(self._get_error_message(self._error_message))
            if self._error_message == [] and not False in self._keys:
                print('Ошибки не обнаружены')
                self._error_file_path.unlink()
                if error == 'verify': sys.exit()
            else:
                sys.exit(f'Найдены ошибки в файле {self._input_file_path}, которые сохранены в файл "{self._error_file_path}"')
        if self._dump_type=='csv':
            with open(self._output_file_path, mode='w', newline='', encoding=self._encoding_output) as csv_file:
                print('CSV создается...')
                writer = csv.writer(csv_file, delimiter=self._delimiter, quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(self._headers)
                for row in tqdm(self._sheet_data.iter_rows(min_row=self._skip_rows+2, min_col=self._skip_cols+1, values_only=True, max_row=self._max_row, max_col=self._max_col)):
                    if any(cell is not None for cell in row):
                        if self._set_empty_str_to_null: row = tuple(x if x != '' else None for x in row)
                        if self._is_strip: row = tuple(x.strip() if isinstance(x, str) else x for x in row)
                        if error != 'ignore': row = self._check_row_validity(row, self._dtypes, None)
                        row = self._insert_extra_values(row)
                        writer.writerow(row)
                print('CSV заполнен')
                print(f'Не забудьте, что используемый разделить: {self._delimiter}')
        elif self._dump_type=='sql':
            batch = list()
            with open(self._output_file_path, mode='w', newline='', encoding=self._encoding_output) as sql_file:
                print('Sql создается...')
                for row in tqdm(self._sheet_data.iter_rows(min_row=self._skip_rows+2, min_col=self._skip_cols+1, values_only=True, max_row=self._max_row, max_col=self._max_col)):
                    if any(cell is not None for cell in row):
                        if self._set_empty_str_to_null: row = tuple(x if x != '' else None for x in row)
                        if self._is_strip: row = tuple(x.strip() if isinstance(x, str) else x for x in row)
                        if error != 'ignore': row = self._check_row_validity(row, self._dtypes, None)
                        row = self._insert_extra_values(row)
                        values = ', '.join([f"'{str(value).replace('\\', '\\\\').replace("'", "''")}'" if value is not None else 'NULL' for value in row])
                        batch.append(f"{values}")
                        if len(batch) >= self._batch_size:
                            sql_file.write(f"INSERT INTO {self._scheme_name}.{self._table_name} ({', '.join(self._headers)}) \n\tVALUES ({'),\n\t\t('.join(batch)});\n")
                            batch = list()
                if batch:
                    batch = [row for row in batch if any(cell is not None for cell in row)]
                    sql_file.write(f"INSERT INTO {self._scheme_name}.{self._table_name} ({', '.join(self._headers)}) \n\tVALUES ({'),\n\t\t('.join(batch)});\n")
                if not (self._scheme_name or self._table_name): print(f'Не забудь заменить scheme_name и\или table_name на свою таблицу приемник')
                print('Sql заполнен')
    def _parse_sql_inserts(self, sql_file, output_encoding='utf-8'):
        tables_data = {}
        with open(sql_file, 'r', encoding=output_encoding) as f:
            sql_content = f.read()
        insert_pattern = re.compile(
            r'INSERT\s+INTO\s+([^\s(]+)\s*\(([^)]+)\)\s*VALUES\s*(.*?)(?=\s*;|$)',
            re.IGNORECASE | re.DOTALL
        )
        for match in insert_pattern.finditer(sql_content):
            table_name = match.group(1).strip('`"[]')
            table_name = table_name.split('.', 1)[1] if '.' in table_name else table_name
            columns = [col.strip(' `"[]') for col in match.group(2).split(',')]
            values_part = match.group(3).strip()
            if table_name not in tables_data:
                tables_data[table_name] = {'columns': columns, 'rows': []}
            batch_values = re.findall(r'\(([^)]+)\)', values_part)
            for values in batch_values:
                row = self._parse_sql_values(values)
                tables_data[table_name]['rows'].append(row)
        return tables_data
    def _parse_sql_values(self, values_str):
        row_values = []
        current_value = []
        in_quotes = False
        quote_char = None
        escape = False
        for char in values_str:
            if escape:
                current_value.append(char)
                escape = False
                continue
            if char == '\\':
                escape = True
                continue
            if char in ('"', "'") and not in_quotes:
                in_quotes = True
                quote_char = char
            elif char == quote_char and in_quotes:
                in_quotes = False
                quote_char = None
            elif char == ',' and not in_quotes:
                row_values.append(self._convert_sql_value(''.join(current_value).strip()))
                current_value = []
                continue
            current_value.append(char)
        if current_value:
            row_values.append(self._convert_sql_value(''.join(current_value).strip()))
        return row_values
    @staticmethod
    def _convert_sql_value(value):
        if value.upper() == 'NULL':
            return None
        elif value.upper() == 'TRUE':
            return True
        elif value.upper() == 'FALSE':
            return False
        elif (value.startswith("'") and value.endswith("'")) or (value.startswith('"') and value.endswith('"')):
            return value[1:-1].replace("\\'", "'").replace('\\"', '"')
        try: return int(value)
        except ValueError:
            try: return float(value)
            except ValueError: return value
    def sql_to_excel(self, input_file_path, encoding_output='utf-8'):
        """
        Преобразование sql файла в эксель(и).
        Args:
            input_file: полный путь к эксель файлу. Обязательный параметр.
            encoding_output: кодировка исходящего файла. По умолчанию 'utf-8'.
        Returns: None (создает эксель с неймингом входящего файла и временной меткой).
        """
        input_file_path = Path(input_file_path)
        tables_data = self._parse_sql_inserts(input_file_path, encoding_output)
        for table_name, data in tables_data.items():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name[:31]
            ws.append(data['columns'])
            for row in tqdm(data['rows']):
                ws.append(row)
            output_file = input_file_path.with_name(f'{table_name}_{datetime.now().strftime('%d%m%y_%H%M%S')}').with_suffix('.xlsx')
            wb.save(output_file)
            print(f"Создан файл: {output_file}")
    def csv_to_excel(self, input_file_path, output_encoding='utf-8', delimiter=None):
        """
        Преобразование csv файла в эксель.
        Args:
            input_file: полный путь к эксель файлу. Обязательный параметр.
            encoding_output: кодировка исходящего файла. По умолчанию 'utf-8'.
            delimiter: тип разделителя для исходящего csv файла. По умолчанию запятая.
        Returns: None (создает эксель с неймингом входящего файла и временной меткой).
        """
        if not delimiter: delimiter = self._delimiter
        input_file_path = Path(input_file_path)
        output_file = input_file_path.with_name(f'{input_file_path.stem}_{datetime.now().strftime('%d%m%y_%H%M%S')}').with_suffix('.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(input_file_path, 'r', encoding=output_encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                ws.append(row)
        wb.save(output_file)
    def _check_reserved_keywords(self, columns):
        clickhouse_reserved_keywords = [
            "ALL", "ALTER", "AND", "ANY", "ARRAY", "AS", "ASCENDING", "ASOF", "BETWEEN",
            "BY", "CASE", "CAST", "CHECK", "CLUSTER", "COLLATE", "COLUMN", "CREATE",
            "CROSS", "DATABASE", "DATE", "DESC", "DESCENDING", "DISTINCT", "DROP",
            "ELSE", "END", "EXISTS", "FINAL", "FIRST", "FORMAT", "FROM", "FULL",
            "GLOBAL", "GROUP", "HAVING", "ID", "IF", "ILIKE", "IN", "INNER", "INSERT",
            "INTERSECT", "INTO", "IS", "JOIN", "KEY", "LEFT", "LIKE", "LIMIT", "NOT",
            "NULL", "OFFSET", "ON", "OPTIMIZE", "OR", "ORDER", "OUTER", "OVER",
            "PARTITION", "PREWHERE", "PRIMARY", "RIGHT", "SAMPLE", "SELECT", "SEMI",
            "SET", "SHOW", "SYNC", "THEN", "TO", "TOP", "TOTALS", "TRAILING", "TRIM",
            "TRUNCATE", "UNION", "UPDATE", "USING", "WHEN", "WHERE", "WINDOW", "WITH"
            ]
        greenplum_reserved_keywords = [
            "ALL", "ANALYSE", "ANALYZE", "AND", "ANY", "ARRAY", "AS", "ASC", "ASYMMETRIC",
            "AUTHORIZATION", "BETWEEN", "BINARY", "BOTH", "CASE", "CAST", "CHECK",
            "COLLATE", "COLUMN", "CONSTRAINT", "CREATE", "CROSS", "CURRENT_DATE",
            "CURRENT_ROLE", "CURRENT_TIME", "CURRENT_TIMESTAMP", "CURRENT_USER",
            "DEFAULT", "DEFERRABLE", "DESC", "DISTINCT", "DO", "ELSE", "END", "EXCEPT",
            "FALSE", "FOR", "FOREIGN", "FREEZE", "FROM", "FULL", "GRANT", "GROUP",
            "HAVING", "ILIKE", "IN", "INITIALLY", "INNER", "INTERSECT", "INTO", "IS",
            "ISNULL", "JOIN", "LEADING", "LEFT", "LIKE", "LIMIT", "LOCALTIME",
            "LOCALTIMESTAMP", "NATURAL", "NOT", "NOTNULL", "NULL", "OFFSET", "ON",
            "ONLY", "OR", "ORDER", "OUTER", "OVERLAPS", "PLACING", "PRIMARY",
            "REFERENCES", "RETURNING", "RIGHT", "SELECT", "SESSION_USER", "SIMILAR",
            "SOME", "SYMMETRIC", "TABLE", "THEN", "TO", "TRAILING", "TRUE", "UNION",
            "UNIQUE", "USER", "USING", "VERBOSE", "WHEN", "WHERE", "WINDOW", "WITH"
            ]
        greenplum_reserved_keywords = list(map(str.lower, greenplum_reserved_keywords))
        clickhouse_reserved_keywords = list(map(str.lower, clickhouse_reserved_keywords))
        unrecommended_columns = []
        for col in columns:
            if self._db_type=='ch' or self._db_type==None:
                if col in clickhouse_reserved_keywords:
                    unrecommended_columns.append(col)
            if col in greenplum_reserved_keywords:
                unrecommended_columns.append(col)
        print(f'Краней не рекомендуется использовать зарезервированные слова в наименовании полей, в частности такие как {'"' + '", "'.join(unrecommended_columns) + '"'}')